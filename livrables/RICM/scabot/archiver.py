"""
Archivage des documents produits sur Drive + journal Excel.
"""
from datetime import date, datetime
from pathlib import Path

import openpyxl
from loguru import logger

from config import settings
from models import DossierTraitement, StatutTraitement


# ─── Journal Excel ───────────────────────────────────────────────────────────

COLONNES_JOURNAL = [
    "date_traitement", "reference_interne", "type_document",
    "fournisseur", "numero_document", "montant_ht", "tva", "montant_ttc",
    "devise", "codes_imputation", "service_demandeur",
    "cpv_retenu", "cpv_libelle", "cpv_score",
    "glb", "destinataire", "statut", "motif_suspension",
    "fichier_original", "bordereau", "glb_fichier",
]


def _ouvrir_ou_creer_journal() -> openpyxl.Workbook:
    chemin = settings.journal_path
    if chemin.exists():
        return openpyxl.load_workbook(chemin)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal SCABOT"
    ws.append(COLONNES_JOURNAL)
    # En-tête en gras
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    return wb


def journaliser(dossier: DossierTraitement) -> None:
    fiche = dossier.fiche
    wb = _ouvrir_ou_creer_journal()
    ws = wb.active

    ligne = {
        "date_traitement":   datetime.now().strftime("%d/%m/%Y %H:%M"),
        "reference_interne": fiche.reference_interne or "",
        "type_document":     fiche.type_document.value if fiche.type_document else "",
        "fournisseur":       fiche.fournisseur_nom or "",
        "numero_document":   fiche.numero_facture_devis or "",
        "montant_ht":        str(fiche.montant_ht or ""),
        "tva":               str(fiche.montant_tva or ""),
        "montant_ttc":       str(fiche.montant_ttc or ""),
        "devise":            fiche.devise,
        "codes_imputation":  " / ".join(fiche.codes_imputation),
        "service_demandeur": fiche.service_demandeur or "",
        "cpv_retenu":        dossier.cpv_retenu.code if dossier.cpv_retenu else "",
        "cpv_libelle":       dossier.cpv_retenu.libelle if dossier.cpv_retenu else "",
        "cpv_score":         f"{dossier.cpv_retenu.score:.0%}" if dossier.cpv_retenu else "",
        "glb":               "Oui" if fiche.glb_requise else "Non",
        "destinataire":      fiche.destinataire_email or settings.email_destinataire_defaut,
        "statut":            dossier.statut.value,
        "motif_suspension":  dossier.motif_suspension or "",
        "fichier_original":  dossier.drive_file_name,
        "bordereau":         Path(dossier.chemin_bordereau).name if dossier.chemin_bordereau else "",
        "glb_fichier":       Path(dossier.chemin_glb).name if dossier.chemin_glb else "",
    }

    ws.append([ligne[col] for col in COLONNES_JOURNAL])
    wb.save(settings.journal_path)
    logger.info(f"Journal mis à jour : {settings.journal_path}")


def deja_traite(file_id: str) -> bool:
    """Vérifie si ce file_id Drive a déjà été traité (idempotence)."""
    chemin = settings.journal_path
    if not chemin.exists():
        return False
    wb = openpyxl.load_workbook(chemin, read_only=True)
    ws = wb.active
    # Cherche file_id dans la colonne "fichier_original" (index 19)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 19 and row[19] == file_id:
            return True
    return False


# ─── Archivage Drive ─────────────────────────────────────────────────────────

def archiver_sur_drive(dossier: DossierTraitement, fichiers: list[Path]) -> None:
    """Upload les documents générés dans le dossier SORTIE Drive."""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
        from google.oauth2.credentials import Credentials

        creds = Credentials.from_authorized_user_file(
            "credentials/token_drive.json",
            scopes=["https://www.googleapis.com/auth/drive.file"]
        )
        service = build("drive", "v3", credentials=creds)

        for f in fichiers:
            if not f.exists():
                continue
            media = MediaFileUpload(str(f), resumable=True)
            meta  = {"name": f.name, "parents": [settings.drive_folder_sortie]}
            service.files().create(body=meta, media_body=media, fields="id").execute()
            logger.info(f"Archivé sur Drive : {f.name}")

    except Exception as e:
        logger.error(f"Archivage Drive échoué : {e}")
